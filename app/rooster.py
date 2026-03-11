"""
ROOSTER - Rule-Based Employee Roster Prediction Logic
Adapted from the original ML RULE BASED ROOSTER.py for web app usage.
"""

import pandas as pd
from datetime import datetime, timedelta
from openpyxl import load_workbook


def clean_roster_excel(file_path: str) -> pd.DataFrame:
    """
    Cleans and reshapes the raw roster Excel file.
    Extracts proper weekday and date columns from headers and flattens data.
    """
    df_raw = pd.read_excel(file_path, header=None)

    employee_info_headers = [
        'Associate ID', 'Associate Name', 'Project ID',
        'Project Allocation End Date', 'Project Manager ID',
        'Start Time', 'End Time', 'City', 'Facility'
    ]

    day_names = df_raw.iloc[0, len(employee_info_headers):].tolist()
    dates_raw = df_raw.iloc[1, len(employee_info_headers):].tolist()

    def normalize_date(date_val):
        try:
            if isinstance(date_val, datetime):
                return date_val.strftime('%m-%d-%Y')
            if isinstance(date_val, (int, float)):
                dt = pd.to_datetime('1899-12-30') + pd.to_timedelta(date_val, unit='D')
                return dt.strftime('%m-%d-%Y')
            return pd.to_datetime(str(date_val), format='%m-%d-%Y', errors='coerce').strftime('%m-%d-%Y')
        except Exception:
            return None

    normalized_dates = [normalize_date(val) for val in dates_raw]
    combined_headers = [
        f"{day}_{date}" if pd.notna(day) and pd.notna(date) else None
        for day, date in zip(day_names, normalized_dates)
    ]

    df_raw.columns = employee_info_headers + combined_headers
    df_raw = df_raw.drop(index=[0, 1])
    df_raw = df_raw[df_raw['Associate Name'].notna()]

    booking_cols = [col for col in df_raw.columns if col not in employee_info_headers and col is not None]
    df_melted = df_raw.melt(
        id_vars=employee_info_headers, value_vars=booking_cols,
        var_name='Day_Date', value_name='Booking_Status'
    )
    df_melted = df_melted[df_melted['Booking_Status'].notna()]
    df_melted[['Weekday', 'Date_Str']] = df_melted['Day_Date'].str.extract(r'(\w+)_([0-9\-]+)')
    df_melted['Date'] = pd.to_datetime(df_melted['Date_Str'], format='%m-%d-%Y', errors='coerce')
    df_melted['Booked'] = df_melted['Booking_Status'].apply(lambda x: 1 if str(x).strip().upper() == 'Y' else 0)

    return df_melted.drop(columns=['Booking_Status', 'Day_Date', 'Date_Str'])


def get_working_days(month: int, year: int, holiday_file: str) -> pd.DataFrame:
    """
    Returns a DataFrame of working days for a given month & year,
    excluding Saturdays, Sundays, and holidays listed in the holiday_file.
    """
    holiday_df = pd.read_excel(holiday_file)
    holiday_df['Date'] = pd.to_datetime(holiday_df['Date'] + f'-{year}', format='%d-%b-%Y', errors='coerce')
    holiday_dates = holiday_df[
        (holiday_df['Kochi'].str.lower().str.strip() == 'holiday') & (holiday_df['Date'].notna())
    ]['Date'].dt.date.tolist()

    start = datetime(year, month, 1)
    end = (start.replace(month=month % 12 + 1, day=1) - timedelta(days=1))
    all_days = [start + timedelta(days=i) for i in range((end - start).days + 1)]

    workdays = [d for d in all_days if d.weekday() < 5 and d.date() not in holiday_dates]
    return pd.DataFrame({'Date': workdays, 'Weekday': [d.strftime('%A') for d in workdays]})


def generate_booking_predictions(
    df_cleaned: pd.DataFrame,
    working_days_df: pd.DataFrame,
    min_days_per_week: int = 3,
    threshold: float = 0.6
) -> pd.DataFrame:
    """
    Generates rule-based booking predictions:
    - Book if historical frequency >= threshold
    - Enforce at least `min_days_per_week` bookings
    """
    df_cleaned = df_cleaned.copy()
    df_cleaned['Weekday'] = df_cleaned['Date'].dt.day_name()

    weekday_freq = (
        df_cleaned.groupby(['Associate ID', 'Associate Name', 'Weekday'])['Booked']
        .mean()
        .reset_index()
        .rename(columns={'Booked': 'Booking_Frequency'})
    )

    unique_employees = weekday_freq[['Associate ID', 'Associate Name']].drop_duplicates()
    schedule = unique_employees.assign(key=1).merge(
        working_days_df.assign(key=1), on='key').drop('key', axis=1)

    schedule = schedule.merge(
        weekday_freq, on=['Associate ID', 'Associate Name', 'Weekday'], how='left'
    )
    schedule['Booking_Frequency'] = schedule['Booking_Frequency'].fillna(0)
    schedule['Predicted'] = (schedule['Booking_Frequency'] >= threshold).astype(int)
    schedule['Week'] = schedule['Date'].apply(lambda d: (d.day - 1) // 7 + 1)

    def enforce_min_bookings(group: pd.DataFrame) -> pd.DataFrame:
        if group['Predicted'].sum() >= min_days_per_week:
            return group
        top_days = group.sort_values('Booking_Frequency', ascending=False).head(min_days_per_week)
        group.loc[top_days.index, 'Predicted'] = 1
        return group

    schedule = schedule.groupby(['Associate ID', 'Week'], group_keys=False).apply(enforce_min_bookings)
    return schedule[['Associate ID', 'Associate Name', 'Date', 'Weekday', 'Predicted']].rename(
        columns={'Predicted': 'Booked'}
    )


def apply_predictions_to_template(
    predicted_df: pd.DataFrame,
    template_path: str,
    output_path: str
) -> None:
    """
    Fills 'Y' into the Excel template file wherever bookings were predicted.
    """
    wb = load_workbook(template_path)
    ws = wb.active
    predicted_df = predicted_df.copy()
    predicted_df['Date'] = pd.to_datetime(predicted_df['Date'], errors='coerce').dt.date

    def normalize_excel_date(cell_val):
        try:
            if isinstance(cell_val, datetime):
                return cell_val.date()
            if isinstance(cell_val, (int, float)):
                return (pd.to_datetime('1899-12-30') + pd.to_timedelta(cell_val, unit='D')).date()
            return datetime.strptime(cell_val[:10], '%m-%d-%Y').date()
        except Exception:
            return None

    date_headers = {
        col: normalize_excel_date(ws.cell(row=2, column=col).value)
        for col in range(10, ws.max_column + 1)
    }

    for row_idx in range(3, ws.max_row + 1):
        associate_id = ws.cell(row=row_idx, column=1).value
        associate_name = str(ws.cell(row=row_idx, column=2).value).strip()
        if not associate_id or not associate_name:
            continue

        predictions = predicted_df[
            (predicted_df['Associate ID'] == associate_id) &
            (predicted_df['Associate Name'].str.strip() == associate_name) &
            (predicted_df['Booked'] == 1)
        ]

        predicted_dates = set(predictions['Date'])

        for col_idx, booking_date in date_headers.items():
            if booking_date in predicted_dates:
                ws.cell(row=row_idx, column=col_idx).value = 'Y'

    wb.save(output_path)


def generate_predictions_to_excel(
    predicted_df: pd.DataFrame,
    output_path: str
) -> None:
    """
    Saves predictions directly to a new Excel file (no template needed).
    """
    output_df = predicted_df.copy()
    output_df['Date'] = pd.to_datetime(output_df['Date']).dt.strftime('%Y-%m-%d')
    output_df['Booking'] = output_df['Booked'].apply(lambda x: 'Y' if x == 1 else '')

    pivot = output_df.pivot_table(
        index=['Associate ID', 'Associate Name'],
        columns='Date',
        values='Booking',
        aggfunc='first'
    ).fillna('')

    pivot.to_excel(output_path)
